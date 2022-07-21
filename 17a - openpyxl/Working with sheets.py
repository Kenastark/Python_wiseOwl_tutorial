
import openpyxl

# create a new workbook
new_book = openpyxl.Workbook()

# will automatically hace Sheet
musical_sheet = new_book.create_sheet("Musicals")

# will now go Sheet / Musicals

# add a worksheet at the start
animations_sheet = new_book.create_sheet("Animations", 0)

# will now go Animation / Sheet / Musical

# add one more sheet in the penultimate position
comedy_sheet = new_book.create_sheet("Comedies", -1)

# will now go Animations / Sheet / Comedies / Musicals

# new_book["Sheet"].title = "Dramas"
new_book.worksheets[1].title = "Dramas"

# write some information to each sheet
animations_sheet["A1"] = "Up"
comedy_sheet["A1"] = "Borat"
new_book["Dramas"]["A1"].value = "Gone with the wind"
musical_sheet["A1"] = "La La Land"

# save my work
new_book.save("/Users/ikenna/Documents/GitHub/Python/17a - openpyxl/Films by genre.xlsx")

new_book.close()


